"""
export.py — формирование Excel-отчётов (ТЗ 14: экспорт результатов)

Возвращает готовый .xlsx в виде bytes для скачивания из дашборда.
Три листа: Сводка, Парк оборудования, Детализация по агрегату.
"""
import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FONT = "Arial"

# заливка по уровням состояния
LEVEL_FILL = {
    "normal":   "E1F5EE", "warning": "FAEEDA",
    "anomaly":  "FAECE7", "critical": "FCEBEB",
}
LEVEL_RU = {"normal": "Норма", "warning": "Наблюдение",
            "anomaly": "Аномалия", "critical": "Критично"}

HEADER_FILL = "185FA5"
THIN = Side(style="thin", color="D9DCE1")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = Font(name=FONT, bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER


def _autofit(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build_report(fleet_rows, summary, econ, recommendations_map, detail_df=None):
    """
    fleet_rows: список dict по агрегатам (из latest_by_asset)
    summary: dict с общими показателями
    econ: dict экономического эффекта
    recommendations_map: fault -> текст рекомендации
    detail_df: (опционально) DataFrame детальной телеметрии одного агрегата
    """
    wb = Workbook()

    # ---------- Лист 1: Сводка ----------
    ws = wb.active
    ws.title = "Сводка"
    ws["A1"] = "Отчёт системы предиктивной диагностики"
    ws["A1"].font = Font(name=FONT, bold=True, size=15, color="1A1D21")
    ws["A2"] = f"Период данных: {summary['start']:%Y-%m-%d} — {summary['end']:%Y-%m-%d}"
    ws["A2"].font = Font(name=FONT, size=10, color="6B7280")
    ws["A3"] = f"Записей: {summary['rows']:,} · Агрегатов: {summary['assets']}"
    ws["A3"].font = Font(name=FONT, size=10, color="6B7280")

    ws["A5"] = "Показатель"; ws["B5"] = "Значение"
    _style_header(ws, 5, 2)
    kpis = [
        ("Всего агрегатов", summary["assets"]),
        ("Критично", summary.get("n_crit", 0)),
        ("Аномалия", summary.get("n_anom", 0)),
        ("Наблюдение", summary.get("n_warn", 0)),
        ("Норма", summary.get("n_norm", 0)),
        ("Средний Health Index", summary.get("avg_hi", "")),
        ("", ""),
        ("Агрегатов с ранним предупреждением", econ["prevented_units"]),
        ("Предотвращённый простой, ч", econ["hours_saved"]),
        ("Валовый эффект, $", econ["gross"]),
        ("Чистый эффект, $", econ["net"]),
    ]
    for i, (k, v) in enumerate(kpis, start=6):
        ws.cell(row=i, column=1, value=k).font = Font(name=FONT, size=10)
        c = ws.cell(row=i, column=2, value=v)
        c.font = Font(name=FONT, size=10, bold=True)
        c.alignment = Alignment(horizontal="right")
    _autofit(ws, [38, 18])

    # ---------- Лист 2: Парк оборудования ----------
    ws2 = wb.create_sheet("Парк оборудования")
    headers = ["Агрегат", "Тип", "Health Index", "Уровень", "Вероятный дефект",
               "Уверенность", "Риск 7д", "Риск 14д", "Риск 30д", "Рекомендация"]
    ws2.append(headers)
    _style_header(ws2, 1, len(headers))

    for s in fleet_rows:
        lvl = s["level"]
        row = [
            s["equipment_id"], s["equipment_type"], round(s["hi"]),
            LEVEL_RU.get(lvl, lvl), s["fault"], f"{s['conf']:.0%}",
            f"{s['risk_7']:.0%}", f"{s['risk_14']:.0%}", f"{s['risk_30']:.0%}",
            recommendations_map.get(s["fault"], ""),
        ]
        ws2.append(row)
        r = ws2.max_row
        fill = LEVEL_FILL.get(lvl, "FFFFFF")
        for c in range(1, len(headers) + 1):
            cell = ws2.cell(row=r, column=c)
            cell.font = Font(name=FONT, size=10)
            cell.border = BORDER
            cell.alignment = Alignment(vertical="center",
                                       wrap_text=(c == len(headers)))
        # подсветка уровня
        ws2.cell(row=r, column=4).fill = PatternFill("solid", fgColor=fill)
    _autofit(ws2, [12, 12, 13, 13, 18, 12, 9, 9, 9, 50])
    ws2.freeze_panes = "A2"

    # ---------- Лист 3: Детализация (опционально) ----------
    if detail_df is not None and len(detail_df):
        ws3 = wb.create_sheet("Детализация")
        cols = ["timestamp", "vibration_rms", "current_a", "temperature",
                "flow_rate", "hi", "anomaly_level", "pred_fault"]
        cols = [c for c in cols if c in detail_df.columns]
        ws3.append(["Время", "Вибрация RMS", "Ток A", "Темп.", "Расход",
                    "Health Index", "Уровень", "Дефект"][:len(cols)])
        _style_header(ws3, 1, len(cols))
        for _, r in detail_df[cols].iterrows():
            vals = list(r.values)
            if hasattr(vals[0], "strftime"):
                vals[0] = vals[0].strftime("%Y-%m-%d %H:%M")
            ws3.append(vals)
        _autofit(ws3, [18, 13, 10, 8, 10, 13, 12, 14])
        ws3.freeze_panes = "A2"

    # в bytes
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


if __name__ == "__main__":
    import sys, os
    sys.path.insert(0, os.path.dirname(__file__))
    from data import load_timeseries, preprocess, data_summary
    from models import DiagnosticEngine
    from diagnostics import RECOMMENDATIONS, economic_impact
    import pandas as pd

    df = preprocess(load_timeseries())
    eng = DiagnosticEngine().fit(df)
    out = eng.predict(df)

    level_rank = {"normal": 0, "warning": 1, "anomaly": 2, "critical": 3}
    fleet = []
    for eid, g in out.groupby("equipment_id"):
        g = g.sort_values("timestamp"); r = g.iloc[-1]
        fleet.append(dict(equipment_id=eid, equipment_type=r["equipment_type"],
                          hi=r["hi"], level=r["anomaly_level"],
                          risk_7=r["risk_7"], risk_14=r["risk_14"], risk_30=r["risk_30"],
                          fault=r["pred_fault"], conf=r["pred_conf"]))
    fleet.sort(key=lambda x: x["hi"])

    summ = data_summary(out)
    summ.update(n_crit=1, n_anom=1, n_warn=1, n_norm=9, avg_hi=82)
    econ = economic_impact(1, 2)
    data = build_report(fleet, summ, econ, RECOMMENDATIONS,
                        detail_df=out[out.equipment_id == fleet[0]["equipment_id"]])
    with open("test_report.xlsx", "wb") as f:
        f.write(data)
    print("Отчёт создан: test_report.xlsx,", len(data), "байт")
